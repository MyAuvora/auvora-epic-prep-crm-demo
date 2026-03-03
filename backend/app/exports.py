"""
Export module for generating downloadable documents and spreadsheets.
Supports CSV, Excel (XLSX), and PDF formats.
"""

import csv
import io
import uuid
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
from dataclasses import dataclass

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Store exports in memory with expiration (in production, use Redis or similar)
exports_store: Dict[str, dict] = {}

# Define allowed datasets and their fields
ALLOWED_DATASETS = {
    "students": {
        "fields": ["student_id", "first_name", "last_name", "grade", "status", "funding_source", 
                   "attendance_present_count", "attendance_absent_count", "attendance_tardy_count",
                   "overall_grade_flag", "overall_risk_flag", "ixl_status_flag"],
        "display_names": {
            "student_id": "ID",
            "first_name": "First Name",
            "last_name": "Last Name",
            "grade": "Grade",
            "status": "Status",
            "funding_source": "Funding",
            "attendance_present_count": "Present",
            "attendance_absent_count": "Absent",
            "attendance_tardy_count": "Tardy",
            "overall_grade_flag": "Grade Status",
            "overall_risk_flag": "Risk Flag",
            "ixl_status_flag": "IXL Status"
        }
    },
    "families": {
        "fields": ["family_id", "family_name", "billing_status", "current_balance", "annual_tuition",
                   "primary_email", "primary_phone"],
        "display_names": {
            "family_id": "ID",
            "family_name": "Family Name",
            "billing_status": "Billing Status",
            "current_balance": "Balance",
            "annual_tuition": "Annual Tuition",
            "primary_email": "Email",
            "primary_phone": "Phone"
        }
    },
    "billing_records": {
        "fields": ["billing_record_id", "family_id", "date", "type", "description", "amount", "source"],
        "display_names": {
            "billing_record_id": "ID",
            "family_id": "Family ID",
            "date": "Date",
            "type": "Type",
            "description": "Description",
            "amount": "Amount",
            "source": "Source"
        }
    },
    "scholarships": {
        "fields": ["scholarship_id", "student_id", "family_id", "scholarship_type", "status",
                   "annual_award_amount", "remaining_balance"],
        "display_names": {
            "scholarship_id": "ID",
            "student_id": "Student ID",
            "family_id": "Family ID",
            "scholarship_type": "Type",
            "status": "Status",
            "annual_award_amount": "Annual Award",
            "remaining_balance": "Remaining"
        }
    },
    "attendance": {
        "fields": ["attendance_id", "student_id", "date", "status", "check_in_time", "check_out_time"],
        "display_names": {
            "attendance_id": "ID",
            "student_id": "Student ID",
            "date": "Date",
            "status": "Status",
            "check_in_time": "Check In",
            "check_out_time": "Check Out"
        }
    },
    "staff": {
        "fields": ["staff_id", "first_name", "last_name", "role", "email", "phone", "status"],
        "display_names": {
            "staff_id": "ID",
            "first_name": "First Name",
            "last_name": "Last Name",
            "role": "Role",
            "email": "Email",
            "phone": "Phone",
            "status": "Status"
        }
    },
    "leads": {
        "fields": ["lead_id", "parent_name", "student_name", "grade_interest", "stage", "source", "created_date"],
        "display_names": {
            "lead_id": "ID",
            "parent_name": "Parent Name",
            "student_name": "Student Name",
            "grade_interest": "Grade Interest",
            "stage": "Stage",
            "source": "Source",
            "created_date": "Created"
        }
    }
}


@dataclass
class ExportSpec:
    """Specification for an export request"""
    dataset: str
    columns: List[str]
    format: str  # csv, xlsx, pdf
    title: str
    filters: Optional[Dict[str, Any]] = None
    sort_by: Optional[str] = None
    sort_order: str = "asc"
    max_rows: int = 1000


def get_field_value(obj: Any, field: str) -> Any:
    """Extract field value from an object, handling enums and nested attributes"""
    if hasattr(obj, field):
        value = getattr(obj, field)
        # Handle enums
        if hasattr(value, 'value'):
            return value.value
        # Handle dates
        if hasattr(value, 'isoformat'):
            return value.isoformat()
        return value
    return ""


def apply_filters(data: List[Any], filters: Dict[str, Any]) -> List[Any]:
    """Apply filters to the data"""
    if not filters:
        return data
    
    filtered = []
    for item in data:
        match = True
        for field, value in filters.items():
            item_value = get_field_value(item, field)
            # Handle string comparison (case-insensitive contains)
            if isinstance(value, str) and isinstance(item_value, str):
                if value.lower() not in item_value.lower():
                    match = False
                    break
            # Handle numeric comparison
            elif isinstance(value, dict):
                if "min" in value and item_value < value["min"]:
                    match = False
                    break
                if "max" in value and item_value > value["max"]:
                    match = False
                    break
            # Handle exact match
            elif item_value != value:
                match = False
                break
        if match:
            filtered.append(item)
    return filtered


def generate_csv(data: List[Dict], columns: List[str], display_names: Dict[str, str]) -> bytes:
    """Generate a CSV file from the data"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    header = [display_names.get(col, col) for col in columns]
    writer.writerow(header)
    
    # Write data rows
    for row in data:
        writer.writerow([row.get(col, "") for col in columns])
    
    return output.getvalue().encode('utf-8')


def generate_xlsx(data: List[Dict], columns: List[str], display_names: Dict[str, str], title: str) -> bytes:
    """Generate an Excel file from the data"""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet names limited to 31 chars
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0A2463", end_color="0A2463", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write header
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=display_names.get(col, col))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    for row_idx, row in enumerate(data, 2):
        for col_idx, col in enumerate(columns, 1):
            value = row.get(col, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            # Format currency columns
            if col in ["amount", "current_balance", "annual_tuition", "annual_award_amount", "remaining_balance"]:
                if isinstance(value, (int, float)):
                    cell.number_format = '$#,##0.00'
    
    # Auto-adjust column widths
    for col_idx, col in enumerate(columns, 1):
        max_length = len(display_names.get(col, col))
        for row in data:
            cell_value = str(row.get(col, ""))
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_pdf(data: List[Dict], columns: List[str], display_names: Dict[str, str], title: str) -> bytes:
    """Generate a PDF file from the data"""
    output = io.BytesIO()
    
    # Use landscape for many columns
    page_size = landscape(letter) if len(columns) > 5 else letter
    doc = SimpleDocTemplate(output, pagesize=page_size, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#0A2463'),
        spaceAfter=12
    )
    elements.append(Paragraph(title, title_style))
    
    # Subtitle with date
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.gray,
        spaceAfter=20
    )
    elements.append(Paragraph(f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", subtitle_style))
    elements.append(Paragraph(f"Total Records: {len(data)}", subtitle_style))
    elements.append(Spacer(1, 12))
    
    # Table data
    table_data = []
    
    # Header row
    header = [display_names.get(col, col) for col in columns]
    table_data.append(header)
    
    # Data rows
    for row in data:
        row_data = []
        for col in columns:
            value = row.get(col, "")
            # Format currency
            if col in ["amount", "current_balance", "annual_tuition", "annual_award_amount", "remaining_balance"]:
                if isinstance(value, (int, float)):
                    value = f"${value:,.2f}"
            row_data.append(str(value) if value is not None else "")
        table_data.append(row_data)
    
    # Calculate column widths
    available_width = page_size[0] - 1*inch
    col_width = available_width / len(columns)
    
    # Create table
    table = Table(table_data, colWidths=[col_width] * len(columns))
    
    # Table style
    table_style = TableStyle([
        # Header style
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0A2463')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        # Data rows style
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        
        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
    ])
    table.setStyle(table_style)
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    return output.getvalue()


def create_export(
    spec: ExportSpec,
    data_context: dict
) -> dict:
    """
    Create an export based on the specification.
    
    Returns:
        Dictionary with export_id, download_url, filename, and expires_at
    """
    # Validate dataset
    if spec.dataset not in ALLOWED_DATASETS:
        return {"error": f"Unknown dataset: {spec.dataset}. Available: {list(ALLOWED_DATASETS.keys())}"}
    
    dataset_config = ALLOWED_DATASETS[spec.dataset]
    
    # Validate columns
    valid_columns = []
    for col in spec.columns:
        if col in dataset_config["fields"]:
            valid_columns.append(col)
    
    if not valid_columns:
        valid_columns = dataset_config["fields"][:6]  # Default to first 6 fields
    
    # Get raw data based on dataset
    raw_data = []
    if spec.dataset == "students":
        raw_data = data_context.get("students", [])
    elif spec.dataset == "families":
        raw_data = data_context.get("families", [])
    elif spec.dataset == "billing_records":
        raw_data = data_context.get("billing_records", [])
    elif spec.dataset == "scholarships":
        raw_data = data_context.get("sufs_scholarships", [])
    elif spec.dataset == "attendance":
        raw_data = data_context.get("attendance_records", [])
    elif spec.dataset == "staff":
        raw_data = data_context.get("staff", [])
    elif spec.dataset == "leads":
        raw_data = data_context.get("leads", [])
    
    # Apply filters
    if spec.filters:
        raw_data = apply_filters(raw_data, spec.filters)
    
    # Sort data
    if spec.sort_by and spec.sort_by in valid_columns:
        reverse = spec.sort_order.lower() == "desc"
        raw_data = sorted(raw_data, key=lambda x: get_field_value(x, spec.sort_by) or "", reverse=reverse)
    
    # Limit rows
    raw_data = raw_data[:spec.max_rows]
    
    # Convert to dictionaries
    data = []
    for item in raw_data:
        row = {}
        for col in valid_columns:
            row[col] = get_field_value(item, col)
        data.append(row)
    
    # Generate file based on format
    display_names = dataset_config["display_names"]
    
    if spec.format == "csv":
        file_bytes = generate_csv(data, valid_columns, display_names)
        mime_type = "text/csv"
        extension = "csv"
    elif spec.format == "xlsx":
        file_bytes = generate_xlsx(data, valid_columns, display_names, spec.title)
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        extension = "xlsx"
    elif spec.format == "pdf":
        file_bytes = generate_pdf(data, valid_columns, display_names, spec.title)
        mime_type = "application/pdf"
        extension = "pdf"
    else:
        return {"error": f"Unsupported format: {spec.format}. Use csv, xlsx, or pdf."}
    
    # Generate export ID and store
    export_id = str(uuid.uuid4())
    filename = f"{spec.title.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{extension}"
    expires_at = datetime.now() + timedelta(minutes=30)
    
    exports_store[export_id] = {
        "file_bytes": file_bytes,
        "filename": filename,
        "mime_type": mime_type,
        "expires_at": expires_at,
        "row_count": len(data)
    }
    
    # Clean up expired exports
    cleanup_expired_exports()
    
    return {
        "export_id": export_id,
        "filename": filename,
        "mime_type": mime_type,
        "row_count": len(data),
        "expires_at": expires_at.isoformat(),
        "download_url": f"/api/exports/{export_id}"
    }


def get_export(export_id: str) -> Optional[dict]:
    """Get an export by ID if it exists and hasn't expired"""
    export = exports_store.get(export_id)
    if not export:
        return None
    
    if datetime.now() > export["expires_at"]:
        del exports_store[export_id]
        return None
    
    return export


def cleanup_expired_exports():
    """Remove expired exports from the store"""
    now = datetime.now()
    expired = [eid for eid, exp in exports_store.items() if now > exp["expires_at"]]
    for eid in expired:
        del exports_store[eid]
